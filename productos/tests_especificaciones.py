"""Carga de especificaciones técnicas desde el Excel del gerente.

El archivo real trae el encabezado en la fila 3, las columnas desde la B, y cada
artículo en un bloque de celdas combinadas con su descripción multilínea en una
sola celda. Nada de eso está garantizado en el próximo archivo, así que el
lector ubica las columnas por nombre.
"""

import io
import os
import tempfile
from contextlib import contextmanager

from django.core.management import CommandError, call_command
from django.test import TestCase

from productos.models import Producto

DESCRIPCION_P62 = (
    'Pallet de superficie cerrada – antideslizante (Reforzado): \n'
    'Codigo P62\n'
    'De medidas externas 1000 x 1200 x 150mm de alto \n'
    'Capacidad de carga estática 2000kgs \n'
    'De nueve patas'
)


@contextmanager
def como_archivo(buf):
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        tmp.write(buf.read())
        ruta = tmp.name
    try:
        yield ruta
    finally:
        os.unlink(ruta)


def excel(hojas, fila_encabezado=3, col_inicial=2,
          titulos=('Codigo', 'Descripcion '), alto_bloque=16,
          hojas_sin_encabezado=()):
    """Arma un .xlsx con el layout del archivo real.

    `hojas` es {nombre: [(codigo, descripcion), ...]}. Cada artículo ocupa un
    bloque combinado de `alto_bloque` filas, como en el original.
    """
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for nombre, filas in hojas.items():
        ws = wb.create_sheet(nombre)
        if titulos and nombre not in hojas_sin_encabezado:
            ws.cell(row=fila_encabezado, column=col_inicial, value=titulos[0])
            ws.cell(row=fila_encabezado, column=col_inicial + 1, value=titulos[1])
        fila = fila_encabezado + 1
        for codigo, descripcion in filas:
            ws.cell(row=fila, column=col_inicial, value=codigo)
            ws.cell(row=fila, column=col_inicial + 1, value=descripcion)
            if alto_bloque > 1:
                for col in (col_inicial, col_inicial + 1):
                    ws.merge_cells(start_row=fila, start_column=col,
                                   end_row=fila + alto_bloque - 1, end_column=col)
            fila += alto_bloque
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class BaseTest(TestCase):
    def importar(self, hojas, **opciones):
        salida = io.StringIO()
        with como_archivo(excel(hojas, **opciones.pop('layout', {}))) as ruta:
            call_command('importar_especificaciones', ruta,
                         stdout=salida, stderr=salida, **opciones)
        return salida.getvalue()

    def articulo(self, codigo='P62', **kwargs):
        return Producto.objects.create(
            codigo=codigo, nombre=f'Artículo {codigo}', **kwargs)


class LecturaTest(BaseTest):
    def test_carga_la_descripcion_multilinea(self):
        self.articulo('P62')
        self.importar({'Pallets': [('P62', DESCRIPCION_P62)]})

        producto = Producto.objects.get(codigo='P62')
        self.assertEqual(len(producto.especificaciones.splitlines()), 5)
        self.assertIn('Capacidad de carga estática 2000kgs', producto.especificaciones)

    def test_guarda_las_lineas_tal_cual_incluido_el_titulo_y_el_codigo(self):
        """Se pidió el texto completo: la primera línea y la del código quedan."""
        self.articulo('P62')
        self.importar({'Pallets': [('P62', DESCRIPCION_P62)]})

        lineas = Producto.objects.get(codigo='P62').especificaciones.splitlines()
        self.assertEqual(lineas[0], 'Pallet de superficie cerrada – antideslizante (Reforzado):')
        self.assertEqual(lineas[1], 'Codigo P62')

    def test_recorta_los_espacios_del_final_de_cada_linea(self):
        self.articulo('P62')
        self.importar({'Pallets': [('P62', 'Una linea   \nOtra linea  ')]})

        self.assertEqual(Producto.objects.get(codigo='P62').especificaciones,
                         'Una linea\nOtra linea')

    def test_recorta_las_lineas_vacias_de_los_extremos(self):
        """El archivo real termina con una línea en blanco."""
        self.articulo('P62')
        self.importar({'Pallets': [('P62', '\nPrimera\nSegunda\n\n')]})

        self.assertEqual(Producto.objects.get(codigo='P62').especificaciones,
                         'Primera\nSegunda')

    def test_conserva_las_lineas_vacias_del_medio(self):
        self.articulo('P62')
        self.importar({'Pallets': [('P62', 'Primera\n\nSegunda')]})

        self.assertEqual(Producto.objects.get(codigo='P62').especificaciones,
                         'Primera\n\nSegunda')

    def test_saltea_las_filas_de_adentro_del_bloque_combinado(self):
        """Las filas combinadas vienen vacías y no son artículos."""
        self.articulo('P62')
        self.articulo('P61')
        salida = self.importar({'Pallets': [('P62', 'Uno'), ('P61', 'Dos')]})

        self.assertIn('bloques en el archivo : 2', salida)

    def test_limpia_el_soft_hyphen_del_codigo(self):
        self.articulo('P62')
        self.importar({'Pallets': [('P\xad62', 'Una especificación')]})

        self.assertEqual(Producto.objects.get(codigo='P62').especificaciones,
                         'Una especificación')


class UbicarColumnasTest(BaseTest):
    """El encabezado se busca por nombre: la posición puede cambiar."""

    def test_lo_encuentra_en_otra_fila_y_otra_columna(self):
        self.articulo('P62')
        self.importar({'Pallets': [('P62', 'Una especificación')]},
                      layout={'fila_encabezado': 7, 'col_inicial': 5})

        self.assertEqual(Producto.objects.get(codigo='P62').especificaciones,
                         'Una especificación')

    def test_acepta_los_titulos_con_acento(self):
        self.articulo('P62')
        self.importar({'Pallets': [('P62', 'Una especificación')]},
                      layout={'titulos': ('Código', 'Descripción')})

        self.assertEqual(Producto.objects.get(codigo='P62').especificaciones,
                         'Una especificación')

    def test_una_hoja_sin_encabezado_se_reporta_y_no_frena_a_las_demas(self):
        self.articulo('P62')
        salida = self.importar({
            'Pallets': [('P62', 'Una especificación')],
            'Notas internas': [('X', 'algo')],
        }, layout={'hojas_sin_encabezado': {'Notas internas'}})

        self.assertIn('hojas sin tabla   : Notas internas', salida)
        self.assertIn('hojas leídas      : 1 de 2', salida)
        self.assertEqual(Producto.objects.get(codigo='P62').especificaciones,
                         'Una especificación')

    def test_un_archivo_sin_tabla_falla_en_lugar_de_no_hacer_nada(self):
        self.articulo('P62')
        with self.assertRaises(CommandError):
            self.importar({'Pallets': [('P62', 'x')]}, layout={'titulos': None})


class VariasHojasTest(BaseTest):
    def test_procesa_todas_las_hojas(self):
        self.articulo('P62')
        self.articulo('R120')
        self.importar({
            'Pallets': [('P62', 'Spec de pallet')],
            'Recipientes': [('R120', 'Spec de recipiente')],
        })

        self.assertEqual(Producto.objects.get(codigo='P62').especificaciones, 'Spec de pallet')
        self.assertEqual(Producto.objects.get(codigo='R120').especificaciones,
                         'Spec de recipiente')

    def test_un_codigo_repetido_gana_el_primero(self):
        self.articulo('P62')
        salida = self.importar({
            'Pallets': [('P62', 'El primero')],
            'Otra': [('P62', 'El repetido')],
        })

        self.assertEqual(Producto.objects.get(codigo='P62').especificaciones, 'El primero')
        self.assertIn('repetido', salida)


class ArticulosQueNoEstanTest(BaseTest):
    def test_no_crea_articulos_nuevos(self):
        """Sin precio ni unidad, inventar un artículo sería peor que avisar."""
        salida = self.importar({'Pallets': [('NO-EXISTE', 'Una especificación')]})

        self.assertFalse(Producto.objects.filter(codigo='NO-EXISTE').exists())
        self.assertIn('no están en el catálogo', salida)
        self.assertIn('NO-EXISTE', salida)

    def test_matchea_el_codigo_sin_importar_mayusculas(self):
        self.articulo('P62')
        self.importar({'Pallets': [('p62', 'Una especificación')]})

        self.assertEqual(Producto.objects.get(codigo='P62').especificaciones,
                         'Una especificación')

    def test_un_bloque_sin_descripcion_se_saltea_y_se_reporta(self):
        self.articulo('P62')
        salida = self.importar({'Pallets': [('P62', '')]})

        self.assertEqual(Producto.objects.get(codigo='P62').especificaciones, '')
        self.assertIn('sin descripción', salida)


class SobrescribirTest(BaseTest):
    def test_avisa_cuando_pisa_algo_ya_cargado(self):
        """Alguien pudo haberlas escrito a mano desde la ficha."""
        self.articulo('P62', especificaciones='Cargado a mano')
        salida = self.importar({'Pallets': [('P62', 'Del Excel')]})

        self.assertEqual(Producto.objects.get(codigo='P62').especificaciones, 'Del Excel')
        self.assertIn('ya tenían especificaciones y se reemplazan', salida)
        self.assertIn('P62', salida)

    def test_no_avisa_cuando_estaban_vacias(self):
        self.articulo('P62')
        salida = self.importar({'Pallets': [('P62', 'Del Excel')]})
        self.assertNotIn('ya tenían especificaciones', salida)

    def test_reimportar_lo_mismo_no_actualiza_nada(self):
        self.articulo('P62')
        self.importar({'Pallets': [('P62', DESCRIPCION_P62)]})
        salida = self.importar({'Pallets': [('P62', DESCRIPCION_P62)]})

        self.assertIn('ya estaban iguales: 1', salida)
        self.assertIn('0 artículo(s) actualizados', salida)

    def test_refresca_la_fecha_de_modificacion(self):
        """bulk_update no dispara auto_now, así que la fecha va a mano."""
        producto = self.articulo('P62')
        antes = producto.updated_at
        self.importar({'Pallets': [('P62', 'Del Excel')]})

        producto.refresh_from_db()
        self.assertGreater(producto.updated_at, antes)


class DryRunTest(BaseTest):
    def test_no_escribe_nada(self):
        self.articulo('P62')
        salida = self.importar({'Pallets': [('P62', 'Del Excel')]}, dry_run=True)

        self.assertEqual(Producto.objects.get(codigo='P62').especificaciones, '')
        self.assertIn('no se escribió nada', salida)

    def test_igual_muestra_el_resumen(self):
        self.articulo('P62', especificaciones='Cargado a mano')
        salida = self.importar({'Pallets': [('P62', 'Del Excel')]}, dry_run=True)

        self.assertIn('se actualizan     : 1', salida)
        self.assertIn('ya tenían especificaciones', salida)


class ArchivoInexistenteTest(TestCase):
    def test_avisa_si_no_encuentra_el_archivo(self):
        with self.assertRaises(CommandError):
            call_command('importar_especificaciones', 'no_existe.xlsx',
                         stdout=io.StringIO())
