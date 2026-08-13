"""Carga las especificaciones técnicas que manda el gerente en un Excel.

El archivo trae, por hoja, una tabla de dos columnas: el código del artículo y
su descripción completa en una sola celda multilínea. Cada artículo ocupa un
bloque de celdas combinadas, así que las filas de adentro del bloque vienen
vacías y se saltean solas.

El encabezado se busca por nombre y no por posición: en el archivo de ejemplo
está en la fila 3 y las columnas arrancan en la B, y eso puede cambiar en el
próximo que manden.
"""

import unicodedata

from django.core.management.base import BaseCommand, CommandError
from django.utils import timezone

from productos.models import Producto

COLUMNA_CODIGO = 'codigo'
COLUMNA_DESCRIPCION = 'descripcion'


def _sin_acentos(texto):
    return ''.join(c for c in unicodedata.normalize('NFD', texto)
                   if unicodedata.category(c) != 'Mn')


def _clave(valor):
    """Normaliza un encabezado para compararlo: sin acentos, sin espacios, minúscula."""
    return _sin_acentos(str(valor or '')).strip().lower()


def _limpiar(valor):
    if valor is None:
        return ''
    return str(valor).replace('­', '').strip()  # soft-hyphen y espacios


def limpiar_descripcion(texto):
    """Deja el texto como vino, sacándole solo lo que no se ve.

    Se respetan las líneas y su orden —incluido el título y la línea del código,
    que es lo que se pidió— pero se recortan los espacios al final de cada línea
    y las líneas vacías de los extremos, que en el Excel sobran y en el PDF no
    imprimen nada.
    """
    if texto is None:
        return ''
    lineas = [linea.rstrip() for linea in str(texto).replace('\r\n', '\n').split('\n')]
    while lineas and not lineas[0]:
        lineas.pop(0)
    while lineas and not lineas[-1]:
        lineas.pop()
    return '\n'.join(lineas)


def ubicar_columnas(ws):
    """(fila_encabezado, col_codigo, col_descripcion), o None si la hoja no las tiene."""
    for numero, fila in enumerate(ws.iter_rows(values_only=True), start=1):
        encontradas = {}
        for indice, celda in enumerate(fila):
            clave = _clave(celda)
            if clave.startswith(COLUMNA_CODIGO):
                encontradas['codigo'] = indice
            elif clave.startswith(COLUMNA_DESCRIPCION):
                encontradas['descripcion'] = indice
        if 'codigo' in encontradas and 'descripcion' in encontradas:
            return numero, encontradas['codigo'], encontradas['descripcion']
        # Un encabezado más abajo de la fila 20 no es un encabezado.
        if numero > 20:
            break
    return None


def leer_hoja(ws):
    """Los bloques (codigo, descripcion) de una hoja, en orden."""
    ubicacion = ubicar_columnas(ws)
    if ubicacion is None:
        return None
    fila_encabezado, col_codigo, col_desc = ubicacion

    bloques = []
    for fila in ws.iter_rows(min_row=fila_encabezado + 1, values_only=True):
        ancho = len(fila)
        codigo = _limpiar(fila[col_codigo]) if col_codigo < ancho else ''
        if not codigo:
            # Fila de adentro de un bloque combinado, o separador.
            continue
        descripcion = limpiar_descripcion(fila[col_desc] if col_desc < ancho else '')
        bloques.append((codigo, descripcion))
    return bloques


class Command(BaseCommand):
    help = 'Carga las especificaciones técnicas de los artículos desde un Excel'

    def add_arguments(self, parser):
        parser.add_argument('archivo', help='Ruta al archivo .xlsx')
        parser.add_argument(
            '--dry-run', action='store_true',
            help='Muestra qué haría y no escribe nada.',
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            raise CommandError('Instalá openpyxl: pip install openpyxl')

        self.stdout.write(f"Leyendo {options['archivo']}...")
        try:
            wb = openpyxl.load_workbook(options['archivo'], data_only=True)
        except FileNotFoundError:
            raise CommandError(f"No encontré el archivo: {options['archivo']}")

        bloques, sin_tabla = self._leer(wb)
        if not bloques:
            raise CommandError(
                'No encontré ninguna fila con código y descripción. '
                'El encabezado tiene que decir "Codigo" y "Descripcion".')

        cambios, iguales, sin_articulo, sin_texto = self._clasificar(bloques)
        self._resumen(wb, bloques, sin_tabla, cambios, iguales, sin_articulo, sin_texto)

        if options['dry_run']:
            self.stdout.write(self.style.WARNING('\n--dry-run: no se escribió nada.'))
            return

        if cambios:
            Producto.objects.bulk_update(
                [p for p, _ in cambios], ['especificaciones', 'updated_at'],
                batch_size=200)
        self.stdout.write(self.style.SUCCESS(
            f'\nListo: {len(cambios)} artículo(s) actualizados.'))

    def _leer(self, wb):
        """Junta los bloques de todas las hojas. El primero gana si se repite."""
        bloques, vistos, sin_tabla = [], set(), []
        for nombre in wb.sheetnames:
            filas = leer_hoja(wb[nombre])
            if filas is None:
                sin_tabla.append(nombre)
                continue
            for codigo, descripcion in filas:
                if codigo.upper() in vistos:
                    self.stderr.write(
                        f'  código repetido en el archivo, gana el primero: {codigo}')
                    continue
                vistos.add(codigo.upper())
                bloques.append((nombre, codigo, descripcion))
        return bloques, sin_tabla

    def _clasificar(self, bloques):
        """Reparte los bloques en: a actualizar, ya iguales, sin artículo, sin texto."""
        # only() a propósito: sin él cada artículo trae su foto binaria al memoria.
        catalogo = {
            p.codigo.upper(): p
            for p in Producto.objects.only('codigo', 'nombre', 'especificaciones',
                                           'updated_at')
        }
        cambios, iguales, sin_articulo, sin_texto = [], [], [], []
        ahora = timezone.now()

        for hoja, codigo, descripcion in bloques:
            producto = catalogo.get(codigo.upper())
            if producto is None:
                sin_articulo.append((hoja, codigo))
                continue
            if not descripcion:
                sin_texto.append((hoja, codigo))
                continue
            if producto.especificaciones == descripcion:
                iguales.append(codigo)
                continue
            tenia = bool(producto.especificaciones.strip())
            producto.especificaciones = descripcion
            # bulk_update no dispara auto_now, así que la fecha va a mano.
            producto.updated_at = ahora
            cambios.append((producto, tenia))
        return cambios, iguales, sin_articulo, sin_texto

    def _resumen(self, wb, bloques, sin_tabla, cambios, iguales, sin_articulo, sin_texto):
        pisados = [p for p, tenia in cambios if tenia]

        self.stdout.write(f'\n  hojas leídas      : {len(wb.sheetnames) - len(sin_tabla)}'
                          f' de {len(wb.sheetnames)}')
        if sin_tabla:
            self.stdout.write(f'  hojas sin tabla   : {", ".join(sin_tabla)}')
        self.stdout.write(f'  bloques en el archivo : {len(bloques)}')
        self.stdout.write(f'  se actualizan     : {len(cambios)}')
        self.stdout.write(f'  ya estaban iguales: {len(iguales)}')

        # Sobrescribir lo que alguien cargó a mano desde la ficha no puede pasar
        # sin avisar: el Excel manda, pero hay que poder ver qué se perdió.
        if pisados:
            self.stdout.write(self.style.WARNING(
                f'\n  {len(pisados)} artículo(s) ya tenían especificaciones y se reemplazan:'))
            for producto in pisados:
                self.stdout.write(f'    {producto.codigo:<14} {producto.nombre[:44]}')

        if sin_articulo:
            self.stdout.write(self.style.WARNING(
                f'\n  {len(sin_articulo)} código(s) del archivo no están en el catálogo:'))
            for hoja, codigo in sin_articulo:
                self.stdout.write(f'    {codigo:<14} (hoja "{hoja}")')

        if sin_texto:
            self.stdout.write(self.style.WARNING(
                f'\n  {len(sin_texto)} código(s) sin descripción, se saltean:'))
            for hoja, codigo in sin_texto:
                self.stdout.write(f'    {codigo:<14} (hoja "{hoja}")')
