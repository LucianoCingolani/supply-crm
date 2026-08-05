"""Importa la lista de clientes exportada del sistema de facturación.

Es idempotente: matchea primero por `id_facturacion` y después por CUIT
normalizado, así se puede volver a correr con una exportación más nueva sin
duplicar nada.
"""

import html
import re
from collections import Counter

from django.core.exceptions import ValidationError
from django.core.management.base import BaseCommand, CommandError
from django.core.validators import validate_email
from django.db import transaction

from clientes.models import Cliente, normalizar_cuit

# Índice de columna en el Excel -> campo del modelo.
# Se ignoran a propósito: 'Tipo Doc.' (un solo valor), 'Domicilio' (4% lleno,
# el que trae la dirección real es el fiscal), la columna sin nombre (vacía) y
# las tres ID_* (códigos internos del facturador, redundantes).
COLUMNAS = {
    0: 'id_facturacion',
    1: 'contacto',
    2: 'dni',
    3: 'razon_social',
    4: 'cuit',
    6: 'domicilio',
    8: 'condicion_fiscal',
    9: 'tipo_factura',
    10: 'provincia',
    11: 'localidad',
    12: 'codigo_postal',
    13: 'email',
    14: 'telefono',
}

# Campos que se completan si vienen vacíos al unir filas duplicadas.
CAMPOS_TEXTO = [c for c in COLUMNAS.values() if c != 'id_facturacion']

PRIMERA_FILA_DE_DATOS = 3   # fila 1 encabezado, fila 2 en blanco

LARGOS = {
    'contacto': 150, 'dni': 15, 'razon_social': 200, 'cuit': 30,
    'domicilio': 200, 'condicion_fiscal': 30, 'tipo_factura': 20,
    'provincia': 60, 'localidad': 100, 'codigo_postal': 10,
    'email': 254, 'telefono': 30,
}


def limpiar(valor):
    """Normaliza una celda: texto, sin entidades HTML, sin espacios de sobra."""
    if valor is None:
        return ''
    # El export trae entidades sin decodificar: 'Ruta Nacional N&#176;5'
    texto = html.unescape(str(valor)).replace('\xa0', ' ')
    return re.sub(r'\s+', ' ', texto).strip()


class Command(BaseCommand):
    help = 'Importa o actualiza clientes desde el Excel del sistema de facturación'

    def add_arguments(self, parser):
        parser.add_argument('archivo', help='Ruta al .xlsx exportado')
        parser.add_argument('--hoja', default=None, help='Nombre de la hoja (por defecto, la primera)')
        parser.add_argument('--dry-run', action='store_true',
                            help='Muestra qué haría sin escribir en la base')

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            raise CommandError('Falta openpyxl: pip install openpyxl')

        wb = openpyxl.load_workbook(options['archivo'], read_only=True, data_only=True)
        try:
            ws = wb[options['hoja']] if options['hoja'] else wb[wb.sheetnames[0]]
            filas = self._leer(ws)
        finally:
            # read_only=True deja el archivo abierto hasta que se cierra a mano.
            wb.close()

        registros, saltadas = self._parsear(filas)
        registros, unidas = self._unir_duplicados(registros)

        creados, actualizados = self._guardar(registros, seco=options['dry_run'])
        self._resumen(len(filas), saltadas, unidas, creados, actualizados, options['dry_run'])

    # ── lectura y parseo ──────────────────────────────────────────

    def _leer(self, ws):
        filas = []
        for cruda in ws.iter_rows(min_row=PRIMERA_FILA_DE_DATOS, values_only=True):
            if any(c is not None and str(c).strip() for c in cruda):
                filas.append(cruda)
        return filas

    def _parsear(self, filas):
        registros, saltadas = [], Counter()

        for cruda in filas:
            datos = {}
            for indice, campo in COLUMNAS.items():
                datos[campo] = limpiar(cruda[indice] if indice < len(cruda) else None)

            # Un valor de 8 dígitos en la columna CUIT es en realidad un DNI.
            digitos = re.sub(r'\D', '', datos['cuit'])
            if len(digitos) == 8:
                datos['dni'] = datos['dni'] or digitos
                datos['cuit'] = ''
            else:
                datos['cuit'] = normalizar_cuit(datos['cuit'])

            # Se necesita algo con qué identificarlo. El DNI entra porque hay
            # filas sin razón social cuyo único dato era el DNI en la columna
            # del CUIT.
            if not any((datos['razon_social'], datos['cuit'], datos['dni'])):
                saltadas['sin_identidad'] += 1
                continue

            if not datos['razon_social']:
                datos['razon_social'] = datos['contacto'] or datos['cuit'] or datos['dni']

            datos['id_facturacion'] = (
                int(datos['id_facturacion']) if datos['id_facturacion'].isdigit() else None
            )
            # Vienen emails sucios ('yolanda,jimenez1967@gmail'). Se descarta el
            # dato, no la fila. Con el validador de Django, no a mano: un chequeo
            # de "tiene arroba" deja pasar justamente esos casos.
            if datos['email']:
                try:
                    validate_email(datos['email'])
                except ValidationError:
                    saltadas['email_invalido'] += 1
                    datos['email'] = ''

            for campo, largo in LARGOS.items():
                if len(datos[campo]) > largo:
                    datos[campo] = datos[campo][:largo]
                    saltadas['truncados'] += 1

            registros.append(datos)

        return registros, saltadas

    def _unir_duplicados(self, registros):
        """Une filas que comparten CUIT. La primera manda; las siguientes solo
        completan campos vacíos, así no se pierde información."""
        por_cuit, resultado, unidas = {}, [], 0

        for datos in registros:
            clave = datos['cuit']
            if clave and clave in por_cuit:
                base = por_cuit[clave]
                for campo in CAMPOS_TEXTO:
                    if not base[campo]:
                        base[campo] = datos[campo]
                unidas += 1
                continue
            if clave:
                por_cuit[clave] = datos
            resultado.append(datos)

        return resultado, unidas

    # ── escritura ─────────────────────────────────────────────────

    def _guardar(self, registros, seco):
        por_id = {c.id_facturacion: c for c in
                  Cliente.objects.exclude(id_facturacion=None).only('id_facturacion')}
        por_cuit = {c.cuit: c for c in
                    Cliente.objects.exclude(cuit='').only('cuit')}

        nuevos, existentes = [], []
        for datos in registros:
            actual = por_id.get(datos['id_facturacion']) or por_cuit.get(datos['cuit'])
            if actual:
                existentes.append((actual.pk, datos))
            else:
                nuevos.append(Cliente(**datos))

        if seco:
            return len(nuevos), len(existentes)

        with transaction.atomic():
            # bulk_create saltea save(), así que el CUIT ya viene normalizado del parseo.
            Cliente.objects.bulk_create(nuevos, batch_size=500)
            for pk, datos in existentes:
                Cliente.objects.filter(pk=pk).update(**datos)

        return len(nuevos), len(existentes)

    def _resumen(self, leidas, saltadas, unidas, creados, actualizados, seco):
        etiqueta = 'SIMULACIÓN (no se escribió nada)' if seco else 'IMPORTACIÓN'
        self.stdout.write(f'\n{etiqueta}')
        self.stdout.write(f'  filas leídas          : {leidas}')
        self.stdout.write(f'  sin nombre ni CUIT    : {saltadas["sin_identidad"]} (salteadas)')
        self.stdout.write(f'  unidas por CUIT igual : {unidas}')
        if saltadas['email_invalido']:
            self.stdout.write(f'  emails descartados    : {saltadas["email_invalido"]}')
        if saltadas['truncados']:
            self.stdout.write(f'  valores truncados     : {saltadas["truncados"]}')
        self.stdout.write(self.style.SUCCESS(f'  clientes nuevos       : {creados}'))
        self.stdout.write(f'  clientes actualizados : {actualizados}')
        if not seco:
            self.stdout.write(f'  total en la base      : {Cliente.objects.count()}')
