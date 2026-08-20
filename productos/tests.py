"""Alta manual de artículos del catálogo e importación masiva."""

import io
import os
import tempfile
from contextlib import contextmanager
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import CommandError, call_command
from django.test import TestCase
from django.urls import reverse

from productos.models import ARS, USD, Categoria, Producto

User = get_user_model()

# Un PNG de 1x1 real: alcanza para verificar que los bytes llegan a la fila.
PNG_1PX = bytes.fromhex(
    '89504e470d0a1a0a0000000d494844520000000100000001080600000'
    '01f15c4890000000a49444154789c6300010000050001'
    '0d0a2db40000000049454e44ae426082'
)


def usuario(role, email):
    return User.objects.create_user(
        email=email, password='x', first_name='Test', last_name='User', role=role,
    )


def datos(**overrides):
    base = {
        'codigo': 'SA-001',
        'nombre': 'Guante de nitrilo azul talle L',
        'unidad_medida': 'PAR',
        'precio': '1500.50',
        'moneda': ARS,
        # El campo es un select: espera la pk de una categoría que exista.
        'categoria': Categoria.desde_nombre('Protección de manos').pk,
    }
    return {**base, **overrides}


class AltaArticuloPermisosTest(TestCase):
    def setUp(self):
        self.url = reverse('productos:create')

    def test_el_empleado_no_entra(self):
        self.client.force_login(usuario(User.EMPLEADO, 'empleado@test.com'))
        respuesta = self.client.get(self.url)
        self.assertRedirects(respuesta, reverse('dashboard'))

    def test_el_empleado_tampoco_puede_crear_por_post(self):
        self.client.force_login(usuario(User.EMPLEADO, 'empleado@test.com'))
        self.client.post(self.url, datos())
        self.assertFalse(Producto.objects.exists())

    def test_el_gerente_entra(self):
        self.client.force_login(usuario(User.GERENTE, 'gerente@test.com'))
        self.assertEqual(self.client.get(self.url).status_code, 200)

    def test_el_admin_entra(self):
        self.client.force_login(usuario(User.ADMIN, 'admin@test.com'))
        self.assertEqual(self.client.get(self.url).status_code, 200)

    def test_el_anonimo_va_al_login(self):
        respuesta = self.client.get(self.url)
        self.assertIn(reverse('accounts:login'), respuesta['Location'])


class AltaArticuloTest(TestCase):
    def setUp(self):
        self.url = reverse('productos:create')
        self.client.force_login(usuario(User.GERENTE, 'gerente@test.com'))

    def test_crea_el_articulo_y_lleva_a_su_ficha(self):
        respuesta = self.client.post(self.url, datos())

        producto = Producto.objects.get(codigo='SA-001')
        self.assertRedirects(respuesta, reverse('productos:detail', args=[producto.pk]))
        self.assertEqual(producto.nombre, 'Guante de nitrilo azul talle L')
        self.assertEqual(producto.unidad_medida, 'PAR')
        self.assertEqual(producto.precio, Decimal('1500.50'))
        self.assertEqual(producto.categoria.nombre, 'Protección de manos')
        self.assertTrue(producto.activo)

    def test_nace_visible_en_el_catalogo(self):
        self.client.post(self.url, datos())
        respuesta = self.client.get(reverse('productos:catalogo'),
                                    {'categoria': 'Protección de manos'})
        self.assertContains(respuesta, 'SA-001')

    def test_guarda_la_imagen_en_la_fila(self):
        imagen = SimpleUploadedFile('foto.png', PNG_1PX, content_type='image/png')
        self.client.post(self.url, datos(imagen=imagen))

        producto = Producto.objects.get(codigo='SA-001')
        self.assertEqual(bytes(producto.foto), PNG_1PX)
        self.assertEqual(producto.foto_tipo, 'image/png')

    def test_la_imagen_es_opcional(self):
        self.client.post(self.url, datos())
        self.assertFalse(Producto.objects.get(codigo='SA-001').foto)

    def test_rechaza_un_archivo_que_no_es_imagen(self):
        archivo = SimpleUploadedFile('lista.pdf', b'%PDF-1.4', content_type='application/pdf')
        respuesta = self.client.post(self.url, datos(imagen=archivo))

        self.assertFalse(Producto.objects.exists())
        self.assertContains(respuesta, 'tiene que ser una imagen')

    def test_rechaza_una_imagen_demasiado_grande(self):
        grande = SimpleUploadedFile('grande.png', b'x' * (5 * 1024 * 1024 + 1), content_type='image/png')
        respuesta = self.client.post(self.url, datos(imagen=grande))

        self.assertFalse(Producto.objects.exists())
        self.assertContains(respuesta, 'no puede superar los 5 MB')

    def test_no_permite_repetir_el_codigo(self):
        Producto.objects.create(codigo='SA-001', nombre='El que ya estaba')
        respuesta = self.client.post(self.url, datos(nombre='El nuevo'))

        self.assertEqual(Producto.objects.filter(codigo='SA-001').count(), 1)
        self.assertEqual(Producto.objects.get(codigo='SA-001').nombre, 'El que ya estaba')
        self.assertEqual(respuesta.status_code, 200)

    def test_acepta_la_coma_como_separador_decimal(self):
        self.client.post(self.url, datos(precio='1500,50'))
        self.assertEqual(Producto.objects.get(codigo='SA-001').precio, Decimal('1500.50'))

    def test_el_precio_puede_quedar_vacio(self):
        self.client.post(self.url, datos(precio=''))
        self.assertIsNone(Producto.objects.get(codigo='SA-001').precio)

    def test_rechaza_un_precio_que_no_es_numero(self):
        respuesta = self.client.post(self.url, datos(precio='mil quinientos'))
        self.assertFalse(Producto.objects.exists())
        self.assertEqual(respuesta.status_code, 200)

    def test_exige_codigo_y_descripcion(self):
        self.client.post(self.url, datos(codigo='', nombre=''))
        self.assertFalse(Producto.objects.exists())

    def test_exige_la_unidad_de_medida(self):
        respuesta = self.client.post(self.url, datos(unidad_medida=''))
        self.assertFalse(Producto.objects.exists())
        self.assertEqual(respuesta.status_code, 200)

    def test_guarda_la_moneda_elegida(self):
        self.client.post(self.url, datos(moneda=USD))
        self.assertEqual(Producto.objects.get(codigo='SA-001').moneda, USD)

    def test_rechaza_una_moneda_inventada(self):
        respuesta = self.client.post(self.url, datos(moneda='BTC'))
        self.assertFalse(Producto.objects.exists())
        self.assertEqual(respuesta.status_code, 200)

    def test_sugiere_las_categorias_que_ya_existen(self):
        Producto.objects.create(codigo='SA-999', nombre='Otro',
                                categoria=Categoria.desde_nombre('Calzado de seguridad'))
        respuesta = self.client.get(self.url)
        self.assertContains(respuesta, 'Calzado de seguridad')

    def test_no_pierde_lo_cargado_cuando_el_formulario_falla(self):
        Producto.objects.create(codigo='SA-001', nombre='El que ya estaba')
        respuesta = self.client.post(self.url, datos(nombre='Casco tipo minero'))
        self.assertContains(respuesta, 'Casco tipo minero')


class EditarDesdeLaFichaTest(TestCase):
    """La ficha del artículo es la pantalla de edición: se entra y se modifica."""

    def setUp(self):
        self.producto = Producto.objects.create(
            codigo='SA-001', nombre='Guante', unidad_medida='PAR',
            precio=Decimal('100'), moneda=ARS,
            categoria=Categoria.desde_nombre('Vieja'),
        )
        self.url = reverse('productos:detail', args=[self.producto.pk])
        self.client.force_login(usuario(User.GERENTE, 'gerente@test.com'))

    def edicion(self, **overrides):
        base = {
            'nombre': 'Guante',
            'unidad_medida': 'PAR',
            'precio': '100',
            'moneda': ARS,
            'categoria': Categoria.desde_nombre('Vieja').pk,
            'especificaciones': '',
        }
        return {**base, **overrides}

    def test_la_ficha_trae_el_formulario_cargado(self):
        respuesta = self.client.get(self.url)
        self.assertContains(respuesta, 'SA-001')
        self.assertContains(respuesta, 'Guardar cambios')

    def test_guarda_y_vuelve_a_la_ficha(self):
        respuesta = self.client.post(self.url, self.edicion(nombre='Guante de nitrilo'))
        self.assertRedirects(respuesta, self.url)
        self.producto.refresh_from_db()
        self.assertEqual(self.producto.nombre, 'Guante de nitrilo')

    def test_cambia_la_categoria(self):
        """Lo que hacía falta para sacar del limbo a lo que se importó."""
        nueva = Categoria.desde_nombre('Protección de manos')
        self.client.post(self.url, self.edicion(categoria=nueva.pk))
        self.producto.refresh_from_db()
        self.assertEqual(self.producto.categoria, nueva)

    def test_cambia_precio_y_moneda_juntos(self):
        self.client.post(self.url, self.edicion(precio='16,81', moneda=USD))
        self.producto.refresh_from_db()
        self.assertEqual(self.producto.precio, Decimal('16.81'))
        self.assertEqual(self.producto.moneda, USD)

    def test_cambia_la_unidad(self):
        self.client.post(self.url, self.edicion(unidad_medida='CAJA'))
        self.producto.refresh_from_db()
        self.assertEqual(self.producto.unidad_medida, 'CAJA')

    def test_cambia_las_especificaciones(self):
        self.client.post(self.url, self.edicion(especificaciones='Talle L\nAzul'))
        self.producto.refresh_from_db()
        self.assertEqual(self.producto.especificaciones, 'Talle L\nAzul')

    def test_el_codigo_no_se_puede_cambiar(self):
        """Es la clave con la que el importador reconoce al artículo."""
        self.client.post(self.url, self.edicion(codigo='OTRO-999'))
        self.producto.refresh_from_db()
        self.assertEqual(self.producto.codigo, 'SA-001')

    def test_una_unidad_inventada_no_guarda_nada(self):
        self.client.post(self.url, self.edicion(unidad_medida='BANANAS', nombre='Cambiado'))
        self.producto.refresh_from_db()
        self.assertEqual(self.producto.unidad_medida, 'PAR')
        self.assertEqual(self.producto.nombre, 'Guante')

    def test_una_moneda_inventada_no_guarda_nada(self):
        self.client.post(self.url, self.edicion(moneda='BTC', nombre='Cambiado'))
        self.producto.refresh_from_db()
        self.assertEqual(self.producto.moneda, ARS)
        self.assertEqual(self.producto.nombre, 'Guante')

    def test_un_precio_invalido_no_guarda_nada(self):
        respuesta = self.client.post(self.url, self.edicion(precio='mil', nombre='Cambiado'))
        self.producto.refresh_from_db()
        self.assertEqual(self.producto.precio, Decimal('100'))
        self.assertEqual(self.producto.nombre, 'Guante')
        self.assertEqual(respuesta.status_code, 200)

    def test_una_unidad_vacia_es_valida_al_editar(self):
        """Lo importado del formato viejo no trae unidad; arreglar el precio de
        una de esas filas no puede exigir además completarla."""
        sin_unidad = Producto.objects.create(codigo='SA-002', nombre='Viejo', moneda=ARS)
        url = reverse('productos:detail', args=[sin_unidad.pk])
        self.client.post(url, {
            'nombre': 'Viejo', 'unidad_medida': '', 'precio': '50',
            'moneda': ARS, 'categoria': '', 'especificaciones': '',
        })
        sin_unidad.refresh_from_db()
        self.assertEqual(sin_unidad.precio, Decimal('50'))

    def test_sugiere_las_categorias_existentes(self):
        Producto.objects.create(codigo='SA-003', nombre='Otro',
                                categoria=Categoria.desde_nombre('Calzado de seguridad'))
        self.assertContains(self.client.get(self.url), 'Calzado de seguridad')


class ImagenDesdeLaFichaTest(TestCase):
    def setUp(self):
        self.producto = Producto.objects.create(
            codigo='SA-001', nombre='Guante', moneda=ARS,
            foto=PNG_1PX, foto_tipo='image/png',
        )
        self.url = reverse('productos:detail', args=[self.producto.pk])
        self.client.force_login(usuario(User.GERENTE, 'gerente@test.com'))

    def base(self, **extra):
        return {'nombre': 'Guante', 'unidad_medida': '', 'precio': '',
                'moneda': ARS, 'categoria': '', 'especificaciones': '', **extra}

    def test_reemplaza_la_imagen(self):
        otra = SimpleUploadedFile('otra.png', b'nuevos-bytes', content_type='image/png')
        self.client.post(self.url, self.base(imagen=otra))
        self.producto.refresh_from_db()
        self.assertEqual(bytes(self.producto.foto), b'nuevos-bytes')

    def test_la_borra(self):
        self.client.post(self.url, self.base(borrar_foto='1'))
        self.producto.refresh_from_db()
        self.assertFalse(self.producto.foto)
        self.assertEqual(self.producto.foto_tipo, '')

    def test_guardar_sin_tocar_la_imagen_no_la_pierde(self):
        self.client.post(self.url, self.base(nombre='Guante nuevo'))
        self.producto.refresh_from_db()
        self.assertEqual(bytes(self.producto.foto), PNG_1PX)


def excel_lista(filas):
    """Arma un .xlsx con el layout del export 'Lista de Articulos'.

    Sin encabezado: la fila 1 ya es un artículo.
    """
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Articulos'
    for fila in filas:
        ws.append(list(fila))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@contextmanager
def como_archivo(buf):
    """call_command hace str() de los posicionales, así que necesita un path."""
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        tmp.write(buf.read())
        ruta = tmp.name
    try:
        yield ruta
    finally:
        os.unlink(ruta)


class ImportarListaTest(TestCase):
    """El export de 4 columnas: código, descripción, unidad, precio."""

    def importar(self, filas, **opciones):
        salida = io.StringIO()
        with como_archivo(excel_lista(filas)) as ruta:
            call_command('importar_articulos', ruta,
                         formato='lista', stdout=salida, stderr=salida, **opciones)
        return salida.getvalue()

    def test_importa_las_cuatro_columnas(self):
        self.importar([('P64', 'Pallet 110 x 110 x 15', 'Un', 47.21)])

        producto = Producto.objects.get(codigo='P64')
        self.assertEqual(producto.nombre, 'Pallet 110 x 110 x 15')
        self.assertEqual(producto.unidad_medida, 'UN')
        self.assertEqual(producto.precio, Decimal('47.21'))

    def test_no_saltea_la_primera_fila(self):
        """A diferencia del export de Enexpro, este no tiene encabezado."""
        self.importar([
            ('A1', 'Primero', 'Un', 10),
            ('A2', 'Segundo', 'Un', 20),
        ])
        self.assertEqual(Producto.objects.count(), 2)
        self.assertTrue(Producto.objects.filter(codigo='A1').exists())

    def test_el_precio_en_cero_queda_sin_precio(self):
        """El export escribe 0 donde no hay precio cargado, y 0 no es un precio."""
        self.importar([('AR4364', 'Armario metálico', 'Un', 0)])
        self.assertIsNone(Producto.objects.get(codigo='AR4364').precio)

    def test_traduce_la_unidad_del_archivo(self):
        self.importar([
            ('A1', 'Por unidad', 'Un', 10),
            ('A2', 'Por kilo', 'KG', 10),
            ('A3', 'Por metro', 'mt.', 10),
        ])
        self.assertEqual(Producto.objects.get(codigo='A1').unidad_medida, 'UN')
        self.assertEqual(Producto.objects.get(codigo='A2').unidad_medida, 'KG')
        self.assertEqual(Producto.objects.get(codigo='A3').unidad_medida, 'M')

    def test_una_unidad_que_no_conoce_queda_vacia(self):
        """Mejor sin unidad que con una equivalencia inventada."""
        self.importar([('A1', 'Algo', 'docena', 10)])
        self.assertEqual(Producto.objects.get(codigo='A1').unidad_medida, '')

    def test_limpia_el_soft_hyphen(self):
        self.importar([('A\xad1', 'Pallet\xad plástico', 'Un', 10)])
        self.assertTrue(Producto.objects.filter(codigo='A1').exists())
        self.assertEqual(Producto.objects.get(codigo='A1').nombre, 'Pallet plástico')

    def test_aplica_la_categoria_que_se_le_pasa(self):
        self.importar([('A1', 'Algo', 'Un', 10)], categoria='Pallets')
        self.assertEqual(Producto.objects.get(codigo='A1').categoria.nombre, 'Pallets')

    def test_aplica_la_moneda_que_se_le_pasa(self):
        self.importar([('A1', 'Algo', 'Un', 10)], moneda=USD)
        self.assertEqual(Producto.objects.get(codigo='A1').moneda, USD)

    def test_por_defecto_importa_en_pesos(self):
        self.importar([('A1', 'Algo', 'Un', 10)])
        self.assertEqual(Producto.objects.get(codigo='A1').moneda, ARS)

    def test_saltea_las_filas_sin_codigo_o_sin_descripcion(self):
        self.importar([
            ('A1', 'Con todo', 'Un', 10),
            ('', 'Sin código', 'Un', 10),
            ('A3', '', 'Un', 10),
        ])
        self.assertEqual(Producto.objects.count(), 1)

    def test_un_codigo_repetido_no_voltea_la_importacion(self):
        """bulk_create explota si el lote trae el mismo código dos veces."""
        salida = self.importar([
            ('A1', 'El primero', 'Un', 10),
            ('A1', 'El repetido', 'Un', 20),
            ('A2', 'Otro', 'Un', 30),
        ])
        self.assertEqual(Producto.objects.count(), 2)
        self.assertEqual(Producto.objects.get(codigo='A1').nombre, 'El primero')
        self.assertIn('repetido', salida)

    def test_reimportar_actualiza_en_lugar_de_duplicar(self):
        self.importar([('A1', 'Precio viejo', 'Un', 10)])
        self.importar([('A1', 'Precio nuevo', 'Un', 99)])

        self.assertEqual(Producto.objects.count(), 1)
        producto = Producto.objects.get(codigo='A1')
        self.assertEqual(producto.nombre, 'Precio nuevo')
        self.assertEqual(producto.precio, Decimal('99'))

    def test_dry_run_no_escribe_nada(self):
        salida = self.importar([('A1', 'Algo', 'Un', 10)], dry_run=True)
        self.assertFalse(Producto.objects.exists())
        self.assertIn('no se escribió nada', salida)

    def test_un_archivo_sin_filas_utiles_falla_en_lugar_de_no_hacer_nada(self):
        with self.assertRaises(CommandError):
            self.importar([('', '', '', None)])

    def test_avisa_si_el_archivo_no_existe(self):
        with self.assertRaises(CommandError):
            call_command('importar_articulos', 'no_existe.xlsx',
                         formato='lista', stdout=io.StringIO())


class ImportarEnexproTest(TestCase):
    """El formato viejo sigue siendo el default y no cambió."""

    def test_saltea_las_dos_filas_de_encabezado(self):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Código', 'Nombre', 'Categoría', 'Subcategoría', 'Precio'])
        ws.append([''] * 5)
        ws.append(['P1', 'Pallet', 'Pallets', 'Plásticos', 1500])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        with como_archivo(buf) as ruta:
            call_command('importar_articulos', ruta, stdout=io.StringIO())

        self.assertEqual(Producto.objects.count(), 1)
        producto = Producto.objects.get(codigo='P1')
        self.assertEqual(producto.categoria.nombre, 'Pallets')
        self.assertEqual(producto.subcategoria, 'Plásticos')


def excel_lista_con_formato(filas):
    """Como excel_lista, pero cada fila trae el number_format de la celda de precio.

    Así es como el export real declara la moneda: 'USD 310,00' y '$ 176.117,50'
    son el mismo número con distinto formato de celda.
    """
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Articulos'
    for codigo, nombre, unidad, precio, formato in filas:
        ws.append([codigo, nombre, unidad, precio])
        if formato:
            ws.cell(row=ws.max_row, column=4).number_format = formato
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


FORMATO_ARS = '"$"\ #,##0.00;\-"$"\ #,##0.00'
FORMATO_USD = '[$USD]\ #,##0.00'


class MonedaDelFormatoTest(TestCase):
    """La moneda no es una columna: está en el formato de la celda de precio."""

    def importar(self, filas, **opciones):
        salida = io.StringIO()
        with como_archivo(excel_lista_con_formato(filas)) as ruta:
            call_command('importar_articulos', ruta,
                         formato='lista', stdout=salida, stderr=salida, **opciones)
        return salida.getvalue()

    def test_lee_pesos_del_formato(self):
        self.importar([('A1', 'Rec. de residuos 14L', 'Un', 176117.50, FORMATO_ARS)])
        producto = Producto.objects.get(codigo='A1')
        self.assertEqual(producto.moneda, ARS)
        self.assertEqual(producto.precio, Decimal('176117.50'))

    def test_lee_dolares_del_formato(self):
        self.importar([('A1', 'Rampa metálica 200L', 'Un', 310, FORMATO_USD)])
        producto = Producto.objects.get(codigo='A1')
        self.assertEqual(producto.moneda, USD)
        self.assertEqual(producto.precio, Decimal('310'))

    def test_cada_fila_lleva_la_suya(self):
        """El mismo archivo mezcla las dos, que es el caso real."""
        self.importar([
            ('A1', 'En pesos', 'Un', 176117.50, FORMATO_ARS),
            ('A2', 'En dólares', 'Un', 310, FORMATO_USD),
        ])
        self.assertEqual(Producto.objects.get(codigo='A1').moneda, ARS)
        self.assertEqual(Producto.objects.get(codigo='A2').moneda, USD)

    def test_el_formato_manda_sobre_la_opcion(self):
        """--moneda es respaldo, no una orden: el archivo sabe más."""
        self.importar([('A1', 'En dólares', 'Un', 310, FORMATO_USD)], moneda=ARS)
        self.assertEqual(Producto.objects.get(codigo='A1').moneda, USD)

    def test_sin_formato_cae_en_el_respaldo(self):
        self.importar([('A1', 'Sin declarar', 'Un', 16.81, None)], moneda=USD)
        self.assertEqual(Producto.objects.get(codigo='A1').moneda, USD)

    def test_avisa_de_las_filas_sin_moneda_que_tienen_precio(self):
        salida = self.importar([('CA279', 'Canasto ventilado', 'Un', 16.81, None)])
        self.assertIn('sin moneda', salida)
        self.assertIn('CA279', salida)

    def test_no_avisa_de_las_que_no_tienen_precio(self):
        """Sin precio la moneda no significa nada; avisar sería solo ruido."""
        salida = self.importar([('A1', 'Sin precio', 'Un', 0, None)])
        self.assertNotIn('sin moneda', salida)

    def test_el_usd_del_formato_no_se_confunde_con_pesos(self):
        """'[$USD]' también contiene un '$': hay que preguntar por USD primero."""
        self.importar([('A1', 'Algo', 'Un', 100, '[$USD]\ #,##0.00;\-[$USD]\ #,##0.00')])
        self.assertEqual(Producto.objects.get(codigo='A1').moneda, USD)


def excel_con_categorias(filas):
    """Layout del export con categorías: la col 0 va vacía y la categoría es un
    encabezado de grupo en la col 1.

    Cada fila es (categoria_o_None, codigo, descripcion, unidad, precio, formato).
    """
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Articulos'
    for categoria, codigo, desc, unidad, precio, formato in filas:
        ws.append([None, categoria, codigo, desc, unidad, precio])
        if formato:
            ws.cell(row=ws.max_row, column=6).number_format = formato
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class ImportarConCategoriasTest(TestCase):
    """El export de 6 columnas donde la categoría encabeza el grupo."""

    def importar(self, filas, **opciones):
        salida = io.StringIO()
        with como_archivo(excel_con_categorias(filas)) as ruta:
            call_command('importar_articulos', ruta, formato='lista-con-categorias',
                         stdout=salida, stderr=salida, **opciones)
        return salida.getvalue()

    def test_lee_las_seis_columnas(self):
        self.importar([('Armarios', '215665', 'ARMARIO', 'Un', 499586.77, FORMATO_ARS)])

        producto = Producto.objects.get(codigo='215665')
        self.assertEqual(producto.categoria.nombre, 'Armarios')
        self.assertEqual(producto.nombre, 'ARMARIO')
        self.assertEqual(producto.unidad_medida, 'UN')
        self.assertEqual(producto.precio, Decimal('499586.77'))
        self.assertEqual(producto.moneda, ARS)

    def test_arrastra_la_categoria_hacia_abajo(self):
        """Solo el primero del grupo la trae; los de abajo la heredan."""
        self.importar([
            ('Armarios', 'A1', 'Armario chico', 'Un', 10, FORMATO_ARS),
            (None, 'A2', 'Armario mediano', 'Un', 20, FORMATO_ARS),
            (None, 'A3', 'Armario grande', 'Un', 30, FORMATO_ARS),
        ])
        for codigo in ('A1', 'A2', 'A3'):
            self.assertEqual(
                Producto.objects.get(codigo=codigo).categoria.nombre, 'Armarios')

    def test_el_grupo_siguiente_corta_el_arrastre(self):
        self.importar([
            ('Armarios', 'A1', 'Armario', 'Un', 10, FORMATO_ARS),
            (None, 'A2', 'Otro armario', 'Un', 20, FORMATO_ARS),
            ('Bandejas', 'B1', 'Bandeja', 'Un', 30, FORMATO_ARS),
            (None, 'B2', 'Otra bandeja', 'Un', 40, FORMATO_ARS),
        ])
        self.assertEqual(Producto.objects.get(codigo='A2').categoria.nombre, 'Armarios')
        self.assertEqual(Producto.objects.get(codigo='B1').categoria.nombre, 'Bandejas')
        self.assertEqual(Producto.objects.get(codigo='B2').categoria.nombre, 'Bandejas')

    def test_le_saca_los_espacios_a_la_categoria(self):
        """El export trae 'Bandejas ' con espacio al final."""
        self.importar([('Bandejas ', 'B1', 'Bandeja', 'Un', 10, FORMATO_ARS)])
        self.assertEqual(Producto.objects.get(codigo='B1').categoria.nombre, 'Bandejas')

    def test_la_moneda_sigue_saliendo_del_formato(self):
        self.importar([
            ('Pallets', 'P1', 'Pallet', 'Un', 49.50, FORMATO_USD),
            (None, 'P2', 'Otro pallet', 'Un', 193542.96, FORMATO_ARS),
        ])
        self.assertEqual(Producto.objects.get(codigo='P1').moneda, USD)
        self.assertEqual(Producto.objects.get(codigo='P2').moneda, ARS)

    def test_las_filas_antes_del_primer_grupo_caen_en_el_respaldo(self):
        salida = self.importar(
            [(None, 'X1', 'Huérfano', 'Un', 10, FORMATO_ARS)],
            categoria='Sin clasificar',
        )
        self.assertEqual(Producto.objects.get(codigo='X1').categoria.nombre,
                         'Sin clasificar')
        self.assertIn('Sin clasificar', salida)

    def test_le_pone_la_categoria_a_lo_que_ya_estaba_sin_clasificar(self):
        """El caso real: los artículos ya están importados, falta clasificarlos."""
        Producto.objects.create(codigo='P1', nombre='Pallet',
                                categoria=Categoria.desde_nombre('Sin clasificar'),
                                precio=Decimal('49.50'), moneda=USD)
        self.importar([('pallets plasticos', 'P1', 'Pallet', 'Un', 49.50, FORMATO_USD)])

        self.assertEqual(Producto.objects.count(), 1)
        producto = Producto.objects.get(codigo='P1')
        self.assertEqual(producto.categoria.nombre, 'pallets plasticos')
        self.assertEqual(producto.precio, Decimal('49.50'))

    def test_no_toca_los_que_no_vienen_en_el_archivo(self):
        """Un export más corto no borra ni desactiva lo que quedó afuera."""
        Producto.objects.create(codigo='VIEJO', nombre='Descontinuado',
                                categoria=Categoria.desde_nombre('Sin clasificar'),
                                precio=Decimal('100'))
        self.importar([('Armarios', 'A1', 'Armario', 'Un', 10, FORMATO_ARS)])

        viejo = Producto.objects.get(codigo='VIEJO')
        self.assertEqual(viejo.categoria.nombre, 'Sin clasificar')
        self.assertTrue(viejo.activo)

    def test_lista_las_categorias_en_el_resumen(self):
        salida = self.importar([
            ('Armarios', 'A1', 'Armario', 'Un', 10, FORMATO_ARS),
            ('Bandejas', 'B1', 'Bandeja', 'Un', 20, FORMATO_ARS),
        ])
        self.assertIn('categorías (2)', salida)
        self.assertIn('Armarios', salida)
        self.assertIn('Bandejas', salida)
