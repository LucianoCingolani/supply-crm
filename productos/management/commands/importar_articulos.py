"""Importa el catálogo desde un Excel del sistema de gestión.

Hay dos exports distintos y ninguno de los dos declara su formato, así que lo
elige quien importa con --formato:

  enexpro  Código | Nombre | Categoría | Subcategoría | Precio, desde la fila 3.
  lista    Código | Descripción | Unidad | Precio, desde la fila 1, sin encabezado.

En el export 'lista' la moneda no es una columna: está en el formato de la
celda de precio ('[$USD] #,##0.00' contra '"$" #,##0.00'), que es lo que hace
que en Excel se lea "USD 310,00" o "$ 176.117,50". El valor crudo es el mismo
número en los dos casos, así que leer solo los valores perdería el dato.
"""

from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError
from django.utils import timezone

from productos.models import ARS, MONEDAS, UNIDADES_MEDIDA, USD, Producto

# Cómo escribe cada export la unidad de medida. La clave va en minúsculas y sin
# puntos; lo que no esté acá queda vacío en lugar de inventar una equivalencia.
UNIDADES_TEXTO = {
    'un': 'UN', 'u': 'UN', 'uni': 'UN', 'unid': 'UN', 'unidad': 'UN',
    'par': 'PAR',
    'caja': 'CAJA', 'cja': 'CAJA',
    'pack': 'PACK',
    'bolsa': 'BOLSA',
    'rollo': 'ROLLO',
    'kg': 'KG', 'kilo': 'KG', 'kilogramo': 'KG',
    'gr': 'GR', 'g': 'GR', 'gramo': 'GR',
    'tn': 'TN', 'ton': 'TN', 'tonelada': 'TN',
    'lt': 'LT', 'l': 'LT', 'litro': 'LT',
    'm': 'M', 'mt': 'M', 'metro': 'M',
    'm2': 'M2', 'm3': 'M3',
}


def _limpiar(valor):
    if valor is None:
        return ''
    return str(valor).replace('­', '').strip()  # elimina soft-hyphen y espacios


def _precio(valor):
    """None cuando no hay precio. El 0 del export significa 'sin precio cargado'."""
    if valor is None:
        return None
    try:
        decimal = Decimal(str(valor))
    except InvalidOperation:
        return None
    return decimal if decimal > 0 else None


def _unidad(texto):
    return UNIDADES_TEXTO.get(_limpiar(texto).lower().rstrip('.'), '')


def moneda_del_formato(number_format, por_defecto):
    """Deduce la moneda del formato de celda. None si el formato no la declara.

    Se pregunta por USD primero: '[$USD]' también contiene un '$'.
    """
    formato = (number_format or '').upper()
    if 'USD' in formato or 'U$S' in formato:
        return USD
    if '$' in formato:
        return ARS
    return por_defecto


def leer_enexpro(ws, opciones):
    for fila in ws.iter_rows(values_only=True, min_row=3):
        codigo, nombre, categoria, subcategoria, precio = (fila + (None,) * 5)[:5]
        yield {
            'codigo': _limpiar(codigo),
            'nombre': _limpiar(nombre),
            'categoria': _limpiar(categoria),
            'subcategoria': _limpiar(subcategoria),
            'unidad_medida': '',
            'precio': _precio(precio),
            'moneda': opciones['moneda'],
            'moneda_declarada': True,
        }


def leer_lista(ws, opciones):
    """El export no trae categoría: la pone quien importa con --categoria."""
    for fila in ws.iter_rows(min_row=1):
        celdas = list(fila) + [None] * 4
        codigo, nombre, unidad, precio = celdas[:4]
        formato = precio.number_format if precio is not None else ''
        declarada = moneda_del_formato(formato, None)
        yield {
            'codigo': _limpiar(codigo.value if codigo is not None else None),
            'nombre': _limpiar(nombre.value if nombre is not None else None),
            'categoria': '',
            'subcategoria': '',
            'unidad_medida': _unidad(unidad.value if unidad is not None else None),
            'precio': _precio(precio.value if precio is not None else None),
            'moneda': declarada or opciones['moneda'],
            'moneda_declarada': declarada is not None,
        }


FORMATOS = {'enexpro': leer_enexpro, 'lista': leer_lista}

ACTUALIZABLES = ['nombre', 'categoria', 'subcategoria', 'unidad_medida',
                 'precio', 'moneda', 'updated_at']


class Command(BaseCommand):
    help = 'Importa o actualiza el catálogo de productos desde un Excel de gestión'

    def add_arguments(self, parser):
        parser.add_argument('archivo', help='Ruta al archivo .xlsx')
        parser.add_argument(
            '--formato', choices=sorted(FORMATOS), default='enexpro',
            help='Layout del archivo (por defecto enexpro).',
        )
        # En 'lista' la moneda sale del formato de cada celda y esto es solo el
        # respaldo para las filas que no la declaran. En 'enexpro' no hay ningún
        # indicio en el archivo, así que rige para todas.
        parser.add_argument(
            '--moneda', choices=[valor for valor, _ in MONEDAS], default=ARS,
            help='Moneda para las filas que no la declaran (por defecto ARS).',
        )
        parser.add_argument(
            '--categoria', default='',
            help='Categoría para todas las filas. Sin ella el artículo no aparece '
                 'en ninguna sección del catálogo, solo buscándolo.',
        )
        parser.add_argument(
            '--dry-run', action='store_true',
            help='Muestra qué haría y no escribe nada.',
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            raise CommandError('Instalá openpyxl: pip install openpyxl')

        self.stdout.write(f"Leyendo {options['archivo']} como formato '{options['formato']}'...")
        try:
            wb = openpyxl.load_workbook(options['archivo'], data_only=True)
        except FileNotFoundError:
            raise CommandError(f"No encontré el archivo: {options['archivo']}")

        filas = list(FORMATOS[options['formato']](wb.active, options))
        objetos, saltados, sin_unidad, sin_moneda = self._armar(filas, options)

        if not objetos:
            raise CommandError('El archivo no tiene ninguna fila utilizable.')

        self._resumen(objetos, saltados, sin_unidad, sin_moneda, options)

        if options['dry_run']:
            self.stdout.write(self.style.WARNING('\n--dry-run: no se escribió nada.'))
            return

        resultado = Producto.objects.bulk_create(
            objetos,
            update_conflicts=True,
            update_fields=ACTUALIZABLES,
            unique_fields=['codigo'],
        )
        self.stdout.write(self.style.SUCCESS(
            f'\nListo: {len(resultado)} productos procesados, {saltados} filas saltadas.'
        ))

    def _armar(self, filas, options):
        objetos, vistos, sin_moneda = [], set(), []
        saltados = sin_unidad = 0
        now = timezone.now()

        for fila in filas:
            if not fila['codigo'] or not fila['nombre']:
                saltados += 1
                continue
            # Un código repetido rompería el bulk_create entero; gana el primero
            # y el resto se informa como salteado.
            if fila['codigo'] in vistos:
                saltados += 1
                self.stderr.write(f"  código repetido, se saltea: {fila['codigo']}")
                continue
            vistos.add(fila['codigo'])
            if not fila['unidad_medida']:
                sin_unidad += 1
            # Solo importa avisar de las que además tienen precio: sin precio, la
            # moneda no significa nada todavía.
            if not fila['moneda_declarada'] and fila['precio'] is not None:
                sin_moneda.append((fila['codigo'], fila['nombre'], fila['precio']))

            objetos.append(Producto(
                codigo=fila['codigo'],
                nombre=fila['nombre'][:300],
                categoria=fila['categoria'] or options['categoria'],
                subcategoria=fila['subcategoria'],
                unidad_medida=fila['unidad_medida'],
                precio=fila['precio'],
                moneda=fila['moneda'],
                updated_at=now,
            ))
        return objetos, saltados, sin_unidad, sin_moneda

    def _resumen(self, objetos, saltados, sin_unidad, sin_moneda, options):
        codigos = [o.codigo for o in objetos]
        existentes = set(
            Producto.objects.filter(codigo__in=codigos).values_list('codigo', flat=True)
        )
        con_precio = sum(1 for o in objetos if o.precio is not None)
        etiquetas = dict(UNIDADES_MEDIDA)

        self.stdout.write(f'\n  filas utilizables : {len(objetos)}')
        self.stdout.write(f'  filas salteadas   : {saltados}')
        self.stdout.write(f'  se crean          : {len(objetos) - len(existentes)}')
        self.stdout.write(f'  se actualizan     : {len(existentes)}')
        self.stdout.write(f'  con precio        : {con_precio}')
        self.stdout.write(f'  sin precio        : {len(objetos) - con_precio}')
        self.stdout.write(f'  categoría         : {options["categoria"] or "(vacía)"}')
        if sin_unidad:
            self.stdout.write(f'  sin unidad recon. : {sin_unidad}')

        monedas = {}
        for o in objetos:
            monedas[o.moneda] = monedas.get(o.moneda, 0) + 1
        self.stdout.write(
            '  monedas           : '
            + ', '.join(f'{m}: {n}' for m, n in sorted(monedas.items()))
        )

        unidades = {}
        for o in objetos:
            unidades[o.unidad_medida] = unidades.get(o.unidad_medida, 0) + 1
        detalle = ', '.join(
            f'{etiquetas.get(u, "sin unidad")}: {n}' for u, n in sorted(unidades.items())
        )
        self.stdout.write(f'  unidades          : {detalle}')

        # Con precio pero sin moneda declarada: quedan en el respaldo, y eso es
        # una suposición que quien importa tiene que poder revisar.
        if sin_moneda:
            self.stdout.write(self.style.WARNING(
                f'\n  {len(sin_moneda)} fila(s) con precio y sin moneda en el archivo. '
                f'Quedan en {options["moneda"]}:'
            ))
            for codigo, nombre, precio in sin_moneda:
                self.stdout.write(f'    {codigo:<14} {nombre[:46]:<48} {precio}')
